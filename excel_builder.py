"""Build the Excel workbook: Data sheet, Summary sheet with charts, Original OCR sheet."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from table_model import TableData, is_number, pivot, totals

# Categorical palette (validated, colour-blind safe order): blue, orange, aqua, yellow,
# magenta, green, violet, red.
SERIES_COLORS = ["2A78D6", "EB6834", "1BAF7A", "EDA100", "E87BA4", "008300", "4A3AA7", "E34948"]
TOTAL_COLOR = "898781"

THIN = Side(style="thin", color="C3C2B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="334155")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="E8E8E4")
ALT_FILL = PatternFill("solid", fgColor="F5F5F3")
FLAG_FILL = PatternFill("solid", fgColor="FFF2A8")
TITLE_FONT = Font(bold=True, size=14)
KPI_FONT = Font(bold=True, size=20, color="0B0B0B")
KPI_LABEL_FONT = Font(size=10, color="52514E")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _autosize(ws, min_w=6, max_w=48):
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            ln = max(len(s) for s in str(cell.value).split("\n"))
            widths[cell.column] = max(widths.get(cell.column, 0), ln)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, w + 2))


def _style_series(chart, n_series: int, total_index: int | None = None):
    for i, s in enumerate(chart.series):
        color = TOTAL_COLOR if i == total_index else SERIES_COLORS[i % len(SERIES_COLORS)]
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = "FFFFFF"  # 1px surface gap between fills


# ---------------------------------------------------------------------------
def _write_data_sheet(ws, td: TableData):
    n_cols = max(td.n_cols, 1)
    r = 1
    if td.title:
        ws.cell(row=1, column=1, value=td.title).font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.row_dimensions[1].height = 26
        r = 3
    header_start = r
    for hr in td.header_rows or [td.columns]:
        for c, v in enumerate(hr, start=1):
            cell = ws.cell(row=r, column=c, value=v or None)
            cell.font, cell.fill, cell.border, cell.alignment = HEADER_FONT, HEADER_FILL, BORDER, CENTER
        r += 1
    for (r1, c1, r2, c2) in td.header_merges:
        try:
            ws.merge_cells(start_row=header_start + r1, start_column=c1 + 1,
                           end_row=header_start + r2, end_column=c2 + 1)
        except ValueError:
            pass
    data_start = r
    for i, row in enumerate(td.rows):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=(v if v != "" else None))
            cell.border = BORDER
            cell.alignment = CENTER if (c - 1) in td.numeric_cols else LEFT
            if is_number(v):
                cell.number_format = "0" if isinstance(v, int) else "#,##0.00"
            if i % 2 == 1:
                cell.fill = ALT_FILL
            if [i, c - 1] in td.flags:
                cell.fill = FLAG_FILL
        r += 1
    data_end = r - 1

    # total row with live formulas (replaces any footer/total row read from the image)
    value_cols = [c for c in td.numeric_cols if c != td.index_col]
    label_col = td.category_col if td.category_col is not None else 0
    if td.rows and value_cols:
        label_written = False
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font, cell.fill, cell.border, cell.alignment = Font(bold=True), TOTAL_FILL, BORDER, CENTER
            if (c - 1) in value_cols:
                col = get_column_letter(c)
                cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
                cell.number_format = "0"
            elif not label_written and (c - 1) not in td.numeric_cols:
                cell.value = "Total"
                label_written = True
        r += 1
    # any extra footer rows from the image, kept as text for reference
    for frow in td.footer_rows[1:]:
        for c, v in enumerate(frow, start=1):
            ws.cell(row=r, column=c, value=(v if v != "" else None)).border = BORDER
        r += 1

    ws.freeze_panes = ws.cell(row=data_start, column=1)
    _autosize(ws)
    for c in td.numeric_cols:
        ws.column_dimensions[get_column_letter(c + 1)].width = max(
            10, min(16, ws.column_dimensions[get_column_letter(c + 1)].width or 10))
    return data_start, data_end


def _write_pivot(ws, top: int, left: int, headers: list[str], rows: list[list], title: str):
    ws.cell(row=top, column=left, value=title).font = Font(bold=True, size=12)
    hr = top + 1
    for j, h in enumerate(headers):
        cell = ws.cell(row=hr, column=left + j, value=h)
        cell.font, cell.fill, cell.border, cell.alignment = HEADER_FONT, HEADER_FILL, BORDER, CENTER
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            cell = ws.cell(row=hr + 1 + i, column=left + j, value=v)
            cell.border = BORDER
            cell.alignment = LEFT if j == 0 else CENTER
            if is_number(v):
                cell.number_format = "0" if isinstance(v, int) else "#,##0.00"
    first, last = hr + 1, hr + len(rows)
    tr = last + 1
    for j in range(len(headers)):
        cell = ws.cell(row=tr, column=left + j)
        cell.font, cell.fill, cell.border, cell.alignment = Font(bold=True), TOTAL_FILL, BORDER, CENTER
        if j == 0:
            cell.value = "Grand Total"
        else:
            col = get_column_letter(left + j)
            cell.value = f"=SUM({col}{first}:{col}{last})"
    return hr, first, last


def _stacked_bar(ws, hr, first, last, left, n_series, title, y_title, x_title):
    chart = BarChart()
    chart.type = "bar"           # horizontal bars, like the reference layout
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 60
    chart.title = title
    chart.style = 10
    chart.y_axis.title = x_title
    chart.x_axis.title = y_title
    chart.y_axis.majorGridlines = None
    chart.legend.position = "b"
    data = Reference(ws, min_col=left + 1, max_col=left + n_series, min_row=hr, max_row=last)
    cats = Reference(ws, min_col=left, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False
    _style_series(chart, n_series)
    chart.height = max(7.5, 1.0 * (last - first + 1) + 3)
    chart.width = 22
    return chart


def _clustered_col(ws, hr, first, last, left, n_series, title):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 80
    chart.title = title
    chart.style = 10
    chart.legend.position = "b"
    chart.y_axis.majorGridlines = None
    data = Reference(ws, min_col=left + 1, max_col=left + n_series, min_row=hr, max_row=last)
    cats = Reference(ws, min_col=left, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False
    _style_series(chart, n_series)
    chart.height = 9
    chart.width = 18
    return chart


def _write_summary_sheet(ws, td: TableData):
    ws.cell(row=1, column=1, value=(td.title or "Summary") + " - Summary").font = TITLE_FONT
    if not td.series_cols or not td.rows:
        ws.cell(row=3, column=1, value="No numeric data detected; nothing to chart.")
        return
    # KPI tiles
    kpis = totals(td)
    col = 1
    for label, val in kpis.items():
        ws.cell(row=3, column=col, value=label).font = KPI_LABEL_FONT
        c = ws.cell(row=4, column=col, value=val)
        c.font, c.number_format = KPI_FONT, "#,##0"
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        col += 2
    ws.row_dimensions[4].height = 30

    by = td.category_col if td.category_col is not None else 0
    headers, rows = pivot(td, by)
    top = 7
    hr, first, last = _write_pivot(ws, top, 1, headers, rows, f"By {td.columns[by]}")
    series_names = [td.columns[c] for c in td.series_cols]
    chart = _stacked_bar(ws, hr, first, last, 1, len(td.series_cols),
                         f"{' / '.join(series_names)} by {td.columns[by]}",
                         td.columns[by], "Count")
    anchor_col = get_column_letter(len(headers) + 2)
    ws.add_chart(chart, f"{anchor_col}{top}")
    next_top = max(last + 4, top + int(chart.height * 2) + 2)

    if td.group_col is not None:
        gheaders, grows = pivot(td, td.group_col)
        ghr, gfirst, glast = _write_pivot(ws, next_top, 1, gheaders, grows, f"By {td.columns[td.group_col]}")
        gchart = _clustered_col(ws, ghr, gfirst, glast, 1, len(td.series_cols),
                                f"{' / '.join(series_names)} by {td.columns[td.group_col]}")
        ws.add_chart(gchart, f"{anchor_col}{next_top}")
    _autosize(ws)
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            14, ws.column_dimensions[get_column_letter(i)].width or 14)


def _write_raw_sheet(ws, td: TableData):
    ws.cell(row=1, column=1, value="Original text as read from the image (before translation)").font = Font(bold=True)
    ws.cell(row=2, column=1, value=f"Engine: {td.engine}   Source language: {td.source_language}")
    r = 4
    for row in td.raw_grid:
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v or None).border = BORDER
        r += 1
    if td.notes:
        r += 1
        ws.cell(row=r, column=1, value="Notes").font = Font(bold=True)
        for n in td.notes:
            r += 1
            ws.cell(row=r, column=1, value=n)
    _autosize(ws, max_w=40)


def build_workbook(td: TableData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    _write_data_sheet(ws, td)
    _write_summary_sheet(wb.create_sheet("Summary"), td)
    _write_raw_sheet(wb.create_sheet("Original OCR"), td)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
